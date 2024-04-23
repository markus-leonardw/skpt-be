from openpyxl import Workbook

import kurikulum.const.predicate as predicate
from kurikulum.enum.prefix import Prefix
from kurikulum.enum.worksheet import Worksheet
from kurikulum.utils.mapper.IRI import IRI


def mapping(wb: Workbook):
    ws = wb[Worksheet.CLO.value]

    min_row = 3
    max_row = wb['Sheet8']['B6'].value + 2

    clo = []
    iri_course_learning_outcome = IRI(Prefix.OBE, 'CourseLearningOutcome')

    for row in ws.iter_rows(min_row=min_row, max_row=max_row, values_only=True):
        iri_clo = IRI(Prefix.OBE, row[0])
        iri_code = IRI(Prefix.STRING, row[0])
        iri_criteria = IRI(Prefix.OBE, row[1])
        iri_desc = IRI(Prefix.OBE, row[2])

        clo.append([iri_clo, predicate.TYPE, iri_course_learning_outcome])
        clo.append([iri_clo, predicate.CODE, iri_code])
        clo.append([iri_clo, predicate.CRITERIA, iri_criteria])
        clo.append([iri_clo, predicate.DESCRIPTION, iri_desc])

    count_clo = wb['Sheet8']['B6'].value
    count_learning_domain = wb['Sheet8']['B10'].value

    ws_learning_domain = wb[Worksheet.LEARNINGDOMAIN.value]
    iri_learning_domains = [IRI(Prefix.OBE, cell[1]) for cell in ws_learning_domain.iter_rows(min_row=2, max_row=count_learning_domain, values_only=True)]

    ws = wb[Worksheet.CLODOMAIN.value]

    for curr_clo in range(3, count_clo + 3):
        iri_clo = IRI(Prefix.OBE, ws.cell(curr_clo, 1).value)
        for curr_ld in range(2, count_learning_domain + 2):
            if ws.cell(curr_clo, curr_ld).value:
                clo.append([iri_clo, predicate.HAS_DOMAIN, iri_learning_domains[curr_ld - 2]])   

    count_plo = wb['Sheet8']['B4'].value
    ws = wb[Worksheet.CLOPLO.value]

    for curr_clo in range(3, count_clo + 3):
        iri_clo = IRI(Prefix.OBE, ws.cell(curr_clo, 1).value)
        for curr_plo in range(2, count_plo + 2):
            iri_plo = IRI(Prefix.OBE, ws.cell(2, curr_plo).value)
            if ws.cell(curr_clo, curr_plo).value:
                clo.append([iri_clo, predicate.CLO_PART_OF_PLO, iri_plo])

    count_splo = wb['Sheet8']['B5'].value
    ws = wb[Worksheet.CLOSPLO.value]

    for curr_clo in range(3, count_clo + 3):
        iri_clo = IRI(Prefix.OBE, ws.cell(curr_clo, 1).value)
        for curr_splo in range(2, count_splo + 2):
            iri_splo = IRI(Prefix.OBE, ws.cell(2, curr_splo).value)
            if ws.cell(curr_clo, curr_splo).value:
                clo.append([iri_clo, predicate.CLO_PART_OF_SPLO, iri_splo])

    return clo