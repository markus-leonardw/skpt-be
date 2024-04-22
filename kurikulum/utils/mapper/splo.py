from openpyxl import Workbook

import kurikulum.const.predicate as predicate
from kurikulum.enum.prefix import Prefix
from kurikulum.enum.worksheet import Worksheet
from kurikulum.utils.mapper.IRI import IRI


def mapping(wb: Workbook):
    ws = wb[Worksheet.SPLO.value]
    
    min_row = 3
    max_row = wb['Sheet8']['B5'].value + 2

    splo = []
    iri_subprogram_learning_outcome = IRI(Prefix.OBE, 'SubProgramLearningOutcome')

    for row in ws.iter_rows(min_row=min_row, max_row=max_row, values_only=True):
        iri_splo = IRI(Prefix.OBE, row[0])
        iri_code = IRI(Prefix.STRING, row[0])
        iri_plo = IRI(Prefix.OBE, row[1])
        iri_desc = IRI(Prefix.STRING, row[2])

        splo.append([iri_splo, predicate.TYPE, iri_subprogram_learning_outcome])
        splo.append([iri_splo, predicate.CODE, iri_code])
        splo.append([iri_splo, predicate.SPLO_PART_OF_PLO, iri_plo])
        splo.append([iri_splo, predicate.DESCRIPTION, iri_desc])

    count_splo = wb['Sheet8']['B5'].value
    count_learning_domain = wb['Sheet8']['B10'].value

    ws_learning_domain = wb[Worksheet.LEARNINGDOMAIN.value]
    iri_learning_domains = [IRI(Prefix.OBE, cell[1]) for cell in ws_learning_domain.iter_rows(min_row=2, max_row=count_learning_domain, values_only=True)]

    ws = wb[Worksheet.SPLODOMAIN.value]

    for curr_splo in range(3, count_splo + 3):
        iri_splo = IRI(Prefix.OBE, ws.cell(curr_splo, 1).value)
        for curr_ld in range(2, count_learning_domain + 2):
            if ws.cell(curr_splo, curr_ld).value:
                splo.append([iri_plo, predicate.HAS_DOMAIN, iri_learning_domains[curr_ld - 2]])   

    return splo