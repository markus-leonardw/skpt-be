from openpyxl import Workbook

import kurikulum.const.predicate as predicate
from kurikulum.enum.prefix import Prefix
from kurikulum.enum.worksheet import Worksheet
from kurikulum.utils.mapper.IRI import IRI


def mapping(wb: Workbook):
    ws = wb[Worksheet.PLO.value]
    
    min_row = 3
    max_row = wb['Sheet8']['B4'].value + 2

    plo = []
    iri_program_learning_outcome = IRI(Prefix.OBE, 'ProgramLearningOutcome')

    for row in ws.iter_rows(min_row=min_row, max_row=max_row, max_col=9, values_only=True):
        # mandatory field
        iri_plo = IRI(Prefix.OBE, row[0])
        iri_code = IRI(Prefix.STRING, row[0])
        iri_desc = IRI(Prefix.STRING, row[1])

        plo.append([iri_plo, predicate.TYPE, iri_program_learning_outcome])
        plo.append([iri_plo, predicate.CODE, iri_code])
        plo.append([iri_plo, predicate.DESCRIPTION, iri_desc])

        # nullable field
        column_to_predicate = {
            2: predicate.KKNI_KNOWLEDGE,
            3: predicate.KKNI_WORKING,
            4: predicate.KKNI_RESPONIBILITY,
            5: predicate.SNDIKTI_ATTITUDE,
            6: predicate.SNDIKTI_KNOWLEDGE,
            7: predicate.SNDIKTI_GENERIC,
            8: predicate.SNDIKTI_SPECIFIC
        }

        # Iterate over the relevant columns and append non-null IRIs to plo
        for column_index, predicate_value in column_to_predicate.items():
            iri = IRI(Prefix.STRING, row[column_index]) if row[column_index] else None
            if iri:
                plo.append([iri_plo, predicate_value, iri])

    count_peo = wb['Sheet8']['B3'].value
    count_plo = wb['Sheet8']['B4'].value
    count_learning_domain = wb['Sheet8']['B10'].value

    ws = wb[Worksheet.PLOPEO.value]

    for curr_plo in range(3, count_plo + 3):
        iri_plo = IRI(Prefix.OBE, ws.cell(curr_plo, 1).value)
        for curr_peo in range(2, count_peo + 2):
            if ws.cell(curr_plo, curr_peo).value:
                iri_peo = IRI(Prefix.OBE, ws.cell(2, curr_peo).value)
                plo.append([iri_plo, predicate.PLO_PART_OF_PEO, iri_peo])

    ws_learning_domain = wb[Worksheet.LEARNINGDOMAIN.value]

    iri_learning_domains = [IRI(Prefix.OBE, cell[1]) for cell in ws_learning_domain.iter_rows(min_row=2, max_row=count_learning_domain, values_only=True)]

    ws = wb[Worksheet.PLODOMAIN.value]

    for curr_plo in range(3, count_plo + 3):
        iri_plo = IRI(Prefix.OBE, ws.cell(curr_plo, 1).value)
        for curr_ld in range(2, count_learning_domain + 2):
            if ws.cell(curr_plo, curr_ld).value:
                plo.append([iri_plo, predicate.HAS_DOMAIN, iri_learning_domains[curr_ld - 2]])   

    return plo
